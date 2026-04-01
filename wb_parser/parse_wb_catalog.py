#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import logging
import time
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


SEARCH_URL = "https://search.wb.ru/exactmatch/ru/common/v5/search"
DEFAULT_QUERY = "пальто из натуральной шерсти"
DEFAULT_DEST = "-1257786"
DEFAULT_CURRENCY = "rub"
DEFAULT_APP_TYPE = "1"
DEFAULT_SPP = "30"
PAGE_SIZE = 100


STOCK_DETAIL_URLS = (
    "https://card.wb.ru/cards/v1/detail",
    "https://card.wb.ru/cards/detail",
)
DEFAULT_REGIONS = (
    "80,38,4,64,83,33,68,70,69,30,86,75,40,1,66,110,22,31,48,71,114"
)

BASKET_HOST_RANGES = (
    (0, 143, 1),
    (144, 287, 2),
    (288, 431, 3),
    (432, 719, 4),
    (720, 1007, 5),
    (1008, 1061, 6),
    (1062, 1115, 7),
    (1116, 1169, 8),
    (1170, 1313, 9),
    (1314, 1601, 10),
    (1602, 1655, 11),
    (1656, 1919, 12),
    (1920, 2045, 13),
    (2046, 2189, 14),
    (2190, 2405, 15),
    (2406, 2621, 16),
    (2622, 2837, 17),
    (2838, 3053, 18),
    (3054, 3269, 19),
    (3270, 3485, 20),
    (3486, 3701, 21),
    (3702, 3917, 22),
    (3918, 4133, 23),
    (4134, 4349, 24),
    (4350, 4565, 25),
    (4566, 4781, 26),
    (4782, 4997, 27),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Парсер каталога Wildberries в XLSX по поисковому запросу."
    )
    parser.add_argument("--query", default=DEFAULT_QUERY, help="Поисковый запрос.")
    parser.add_argument(
        "--output-dir",
        default="wb_parser/output",
        help="Папка, куда будут сохранены XLSX-файлы.",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=8.0,
        help="Пауза между запросами к страницам поиска в секундах.",
    )
    parser.add_argument(
        "--max-pages",
        type=int,
        default=50,
        help="Лимит страниц поиска для защиты от бесконечного цикла.",
    )
    parser.add_argument(
        "--log-level",
        default="INFO",
        choices=("DEBUG", "INFO", "WARNING", "ERROR"),
        help="Уровень логирования.",
    )
    return parser.parse_args()


def build_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/123.0.0.0 Safari/537.36"
            ),
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        }
    )
    return session


def fetch_json(
    session: requests.Session,
    url: str,
    *,
    params: Optional[Dict[str, object]] = None,
    timeout: int = 30,
    retries: int = 5,
    retry_sleep: float = 5.0,
    expected_statuses: Sequence[int] = (200,),
) -> Optional[dict]:
    for attempt in range(1, retries + 1):
        try:
            response = session.get(url, params=params, timeout=timeout)
        except requests.RequestException as exc:
            logging.warning("Request failed for %s: %s", url, exc)
            if attempt == retries:
                return None
            time.sleep(retry_sleep * attempt)
            continue

        if response.status_code in expected_statuses:
            try:
                return response.json()
            except ValueError:
                logging.warning("Invalid JSON from %s", response.url)
                return None

        logging.warning(
            "Unexpected status %s for %s", response.status_code, response.url
        )
        if response.status_code == 429 and attempt < retries:
            time.sleep(retry_sleep * attempt)
            continue
        if attempt == retries:
            return None
        time.sleep(retry_sleep * attempt)
    return None


def fetch_search_page(
    session: requests.Session, query: str, page: int, *, page_size: int = PAGE_SIZE
) -> List[dict]:
    params = {
        "ab_testing": "false",
        "appType": DEFAULT_APP_TYPE,
        "curr": DEFAULT_CURRENCY,
        "dest": DEFAULT_DEST,
        "query": query,
        "resultset": "catalog",
        "sort": "popular",
        "spp": DEFAULT_SPP,
        "limit": page_size,
        "page": page,
    }
    payload = fetch_json(session, SEARCH_URL, params=params)
    if not payload:
        return []
    return payload.get("products", []) or []


def get_basket_host_by_volume(vol: int) -> str:
    for start, end, basket_number in BASKET_HOST_RANGES:
        if start <= vol <= end:
            return f"basket-{basket_number:02d}.wbbasket.ru"
    return "basket-01.wbbasket.ru"


def build_card_json_url(host: str, nm_id: int) -> str:
    vol = nm_id // 100000
    part = nm_id // 1000
    return f"https://{host}/vol{vol}/part{part}/{nm_id}/info/ru/card.json"


def resolve_card_payload(
    session: requests.Session, nm_id: int, host_cache: Dict[int, str]
) -> tuple[Optional[dict], Optional[str]]:
    vol = nm_id // 100000
    candidate_hosts = []
    cached = host_cache.get(vol)
    if cached:
        candidate_hosts.append(cached)
    guessed = get_basket_host_by_volume(vol)
    if guessed not in candidate_hosts:
        candidate_hosts.append(guessed)
    candidate_hosts.extend(
        f"basket-{index:02d}.wbbasket.ru"
        for index in range(1, 31)
        if f"basket-{index:02d}.wbbasket.ru" not in candidate_hosts
    )

    for host in candidate_hosts:
        payload = fetch_json(
            session,
            build_card_json_url(host, nm_id),
            retries=2,
            retry_sleep=1.0,
        )
        if payload:
            host_cache[vol] = host
            return payload, host
    return None, None


def fetch_stock_count(session: requests.Session, nm_id: int) -> Optional[int]:
    params = {
        "appType": DEFAULT_APP_TYPE,
        "curr": DEFAULT_CURRENCY,
        "dest": DEFAULT_DEST,
        "regions": DEFAULT_REGIONS,
        "spp": DEFAULT_SPP,
        "nm": nm_id,
    }
    for url in STOCK_DETAIL_URLS:
        payload = fetch_json(
            session,
            url,
            params=params,
            retries=2,
            retry_sleep=1.0,
        )
        if not payload:
            continue
        products = (
            payload.get("data", {}).get("products", [])
            if isinstance(payload, dict)
            else []
        )
        if not products:
            continue
        total_qty = 0
        for size in products[0].get("sizes", []) or []:
            for stock in size.get("stocks", []) or []:
                total_qty += int(stock.get("qty") or 0)
        return total_qty
    return None


def price_to_rubkopecks(price_value: Optional[int]) -> Optional[float]:
    if price_value is None:
        return None
    return round(price_value / 100, 2)


def choose_price(product: dict) -> Optional[float]:
    prices = []
    for size in product.get("sizes", []) or []:
        price_block = size.get("price") or {}
        if price_block.get("product") is not None:
            prices.append(price_block["product"])
    if not prices:
        return None
    return price_to_rubkopecks(min(prices))


def unique_join(values: Iterable[object]) -> str:
    seen = set()
    result: List[str] = []
    for value in values:
        if value in (None, ""):
            continue
        text = str(value).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return ", ".join(result)


def format_characteristics(grouped_options: object) -> str:
    if not grouped_options:
        return ""
    return json.dumps(grouped_options, ensure_ascii=False, indent=2)


def extract_country(card_payload: dict) -> str:
    for option in card_payload.get("options", []) or []:
        if str(option.get("name")).strip().lower() == "страна производства":
            return str(option.get("value") or "").strip()
    return ""


def build_image_urls(host: Optional[str], nm_id: int, media: dict) -> str:
    if not host:
        return ""
    photo_count = int((media or {}).get("photo_count") or 0)
    if photo_count <= 0:
        return ""
    vol = nm_id // 100000
    part = nm_id // 1000
    urls = [
        f"https://{host}/vol{vol}/part{part}/{nm_id}/images/big/{index}.webp"
        for index in range(1, photo_count + 1)
    ]
    return ", ".join(urls)


def normalize_product_record(
    product: dict, card_payload: dict, card_host: Optional[str], stock_count: Optional[int]
) -> dict:
    nm_id = int(product["id"])
    supplier_id = product.get("supplierId")
    seller_name = str(product.get("supplier") or "").strip()
    review_rating = product.get("reviewRating")
    if review_rating is None:
        review_rating = product.get("nmReviewRating")
    if review_rating is None:
        review_rating = product.get("rating")

    record = {
        "Ссылка на товар": f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
        "Артикул": nm_id,
        "Название": card_payload.get("imt_name") or product.get("name") or "",
        "Цена": choose_price(product),
        "Описание": card_payload.get("description") or "",
        "Ссылки на изображения через запятую": build_image_urls(
            card_host, nm_id, card_payload.get("media") or {}
        ),
        "Все характеристики с сохранением структуры": format_characteristics(
            card_payload.get("grouped_options")
        ),
        "Название селлера": seller_name,
        "Ссылка на селлера": (
            f"https://www.wildberries.ru/seller/{supplier_id}" if supplier_id else ""
        ),
        "Размеры товара через запятую": unique_join(
            size.get("name") or size.get("origName")
            for size in product.get("sizes", []) or []
        ),
        "Остатки по товару (число)": stock_count,
        "Рейтинг": review_rating,
        "Количество отзывов": product.get("feedbacks"),
        "_country": extract_country(card_payload),
    }
    return record


def adjust_worksheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 26,
        "B": 14,
        "C": 32,
        "D": 14,
        "E": 70,
        "F": 70,
        "G": 80,
        "H": 24,
        "I": 28,
        "J": 24,
        "K": 22,
        "L": 12,
        "M": 18,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def save_xlsx(path: Path, records: List[dict]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Каталог"

    columns = [
        "Ссылка на товар",
        "Артикул",
        "Название",
        "Цена",
        "Описание",
        "Ссылки на изображения через запятую",
        "Все характеристики с сохранением структуры",
        "Название селлера",
        "Ссылка на селлера",
        "Размеры товара через запятую",
        "Остатки по товару (число)",
        "Рейтинг",
        "Количество отзывов",
    ]
    worksheet.append(columns)

    for record in records:
        worksheet.append([record.get(column) for column in columns])

    adjust_worksheet(worksheet)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def filter_records(records: List[dict]) -> List[dict]:
    filtered = []
    for record in records:
        rating = record.get("Рейтинг")
        price = record.get("Цена")
        country = str(record.get("_country") or "").strip().lower()
        if rating is None or price is None:
            continue
        if float(rating) < 4.5:
            continue
        if float(price) > 10000:
            continue
        if country != "россия":
            continue
        filtered.append(record)
    return filtered


def collect_catalog(
    session: requests.Session,
    query: str,
    *,
    max_pages: int,
    delay: float,
) -> List[dict]:
    host_cache: Dict[int, str] = {}
    records: List[dict] = []
    seen_ids = set()

    for page in range(1, max_pages + 1):
        logging.info("Fetching search page %s", page)
        products = fetch_search_page(session, query, page)
        if not products:
            logging.info("No products returned on page %s, stopping.", page)
            break

        page_new_ids = 0
        for index, product in enumerate(products, start=1):
            nm_id = int(product["id"])
            if nm_id in seen_ids:
                continue
            seen_ids.add(nm_id)
            page_new_ids += 1

            logging.info(
                "Processing product %s/%s on page %s: %s",
                index,
                len(products),
                page,
                nm_id,
            )
            card_payload, card_host = resolve_card_payload(session, nm_id, host_cache)
            if not card_payload:
                logging.warning("Skipping nm_id=%s because card.json was not found", nm_id)
                continue

            stock_count = fetch_stock_count(session, nm_id)
            records.append(
                normalize_product_record(product, card_payload, card_host, stock_count)
            )
            time.sleep(0.5)

        if page_new_ids == 0:
            logging.info("Page %s contains only already seen ids, stopping.", page)
            break
        if len(products) < PAGE_SIZE:
            logging.info("Page %s returned %s products, stopping.", page, len(products))
            break
        time.sleep(delay)

    return records


def main() -> int:
    args = parse_args()
    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s %(levelname)s %(message)s",
    )

    session = build_session()
    records = collect_catalog(
        session,
        args.query,
        max_pages=args.max_pages,
        delay=args.delay,
    )

    output_dir = Path(args.output_dir)
    full_catalog_path = output_dir / "wildberries_catalog_full.xlsx"
    filtered_catalog_path = output_dir / "wildberries_catalog_filtered.xlsx"

    save_xlsx(full_catalog_path, records)
    save_xlsx(filtered_catalog_path, filter_records(records))

    logging.info("Saved full catalog to %s", full_catalog_path)
    logging.info("Saved filtered catalog to %s", filtered_catalog_path)
    logging.info("Collected %s products", len(records))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
